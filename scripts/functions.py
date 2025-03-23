from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import ElementClickInterceptedException, StaleElementReferenceException, TimeoutException

import pandas as pd
import os
import time
from openpyxl import load_workbook

def login(driver, username, password):
    """Login to Alma"""
    username_field = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'username')))
    username_field.send_keys(username)

    password_field = driver.find_element(By.ID, 'password')
    password_field.send_keys(password)

    password_field.submit()
    return driver

def safe_find_element(driver, by, value, retries=3):
    """Find element with retries, and refresh the page if retries fail."""
    for attempt in range(retries):
        try:
            return WebDriverWait(driver, 10).until(EC.visibility_of_element_located((by, value)))
        except StaleElementReferenceException:
            print(f"Retry {attempt + 1} of {retries}: Element stale, retrying...")
            time.sleep(2)
            if attempt == retries - 1:
                print("Too many stale element errors. Refreshing page...")
                driver.refresh()  # Refresh the page on final attempt
                time.sleep(5)
    print(f"Failed to locate element: {value} after {retries} retries.")
    return None
def safe_find_element_text(driver, by, value, retries=3):
    """Find element with retries for StaleElementReferenceException"""

    for attempt in range(retries):
        try:
            element =  WebDriverWait(driver, 20).until(EC.visibility_of_element_located((by, value)))
            return element.text
        except StaleElementReferenceException:
            print(f"Retry {attempt + 1} of {retries}: Element stale, retrying...")
            time.sleep(2)
            if attempt == retries - 1:
                print("Too many stale element errors. Refreshing page...")
                driver.refresh()  # Refresh the page on final attempt
                time.sleep(5)
    print(f"Failed to locate element: {value} after {retries} retries.")
    return None  # Return None instead of throwing an exception
def click_element_with_retry(driver, by, value, retries=3, wait_time=10):
    """Click an element with retry to handle stale elements dynamically."""
    for attempt in range(retries):
        try:
            # Wait for element to be present
            element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((by, value)))

            # Wait until element is clickable
            element = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((by, value)))
             # Click the element
            element.click()
            return
        except ElementClickInterceptedException:
            print(f"Attempt {attempt + 1} of {retries}: Click intercepted by overlay, handling...")

            # Check if an overlay is blocking and remove it
            try:
                overlay = driver.find_element(By.CLASS_NAME, "mask")  # Adjust class name if needed
                driver.execute_script("arguments[0].remove();", overlay)  # Remove the overlay
                print("Overlay detected and removed.")
            except:
                print("No overlay found.")

            time.sleep(2)  # Give time for changes before retrying
        except StaleElementReferenceException:
            print(f"Retry {attempt + 1} of {retries}: Element stale, retrying...")
            time.sleep(2)
            if attempt == retries - 1:
                print("Too many stale element errors. Refreshing page...")
                driver.refresh()  # Refresh the page on final attempt
                time.sleep(5)
                print(f"Failed to locate element: {value} after {retries} retries.")
    return None  # Return None instead of throwing an exception
def get_table_html_with_retry(driver, by, value, retries=3):
    """Retries getting table HTML to handle stale element issues."""
    for attempt in range(retries):
        try:
            table_element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((by, value)))
            return table_element.get_attribute('outerHTML')
        except StaleElementReferenceException:
            print(f"Attempt {attempt + 1} of {retries}: Table element stale, retrying...")
            time.sleep(2)
    print(f"Failed to locate element: {value} after {retries} retries.")
    return None  # Return None instead of throwing an exception
def send_keys_with_retry(driver, by, value, text, retries=3, wait_time=10):
    """Send keys to an element with retry to handle stale element issues."""
    for attempt in range(retries):
        try:
            element = WebDriverWait(driver, wait_time).until(
                EC.element_to_be_clickable((by, value))
            )
            element.clear()
            element.send_keys(text)
            return
        except:
            print(f"Attempt {attempt + 1} of {retries}: Stale element reference, retrying...")
            time.sleep(2)
    print(f"Failed to locate element: {value} after {retries} retries.")
    return None  # Return None instead of throwing an exception
import pandas as pd
from openpyxl import load_workbook

def append_to_excel(file_path, buffer):
    """Append buffer data to Excel file in batches or force flush when needed"""

    buffer_df = pd.DataFrame(buffer)
    print(file_path)
    # Ensure existing data is loaded if file exists
    if os.path.exists(file_path):
        if os.path.exists(file_path):
            with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                buffer_df.to_excel(writer, index=False, header=False, startrow=writer.sheets['Sheet1'].max_row)
    else:
        buffer_df.to_excel(file_path, index=False)

    print(f"🔹 Wrote {len(buffer_df)} records to {file_path} (Total size: {os.path.getsize(file_path) / 1024:.2f} KB)")

    buffer.clear()  # Clear buffer after writing

