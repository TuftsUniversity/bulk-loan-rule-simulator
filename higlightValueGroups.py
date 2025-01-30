import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def highlight_unique_values(file_path, output_path):
    # Load the spreadsheet into a pandas DataFrame
    df = pd.read_excel(file_path, engine='openpyxl')

    # Ensure column G exists
    if len(df.columns) < 7:  # Column G is the 8th column (0-indexed)
        raise ValueError("Column G does not exist in the spreadsheet.")

    df = df.sort_values(by=['TOU (Loan)'])
    # Get unique values in column G
    unique_values = df.iloc[:, 6].dropna().unique()  # Column H (0-indexed)
    color_map = {}

    # Generate unique colors for each unique value in Column H
    for i, value in enumerate(unique_values):
        # Generate color codes in ARGB format (8-character hex string)
        red = (100 + (i * 50) % 256) % 256
        green = (150 + (i * 30) % 256) % 256
        blue = (200 + (i * 70) % 256) % 256
        color_map[value] = f"FF{red:02X}{green:02X}{blue:02X}"

    # Load workbook and active sheet
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # Iterate through column H and apply fill
    for row in range(2, sheet.max_row + 1):  # Skip header (row 1)
        cell = sheet[f'G{row}']
        value = cell.value
        if value in color_map:
            fill = PatternFill(start_color=color_map[value], end_color=color_map[value], fill_type="solid")
            cell.fill = fill

    # Save the updated workbook
    workbook.save(output_path)
    print(f"File saved with highlighted column H: {output_path}")

# Usage
input_file = "Bulk_Checkout_Request_Results.xlsx"  # Path to your input file
output_file = "Bulk_Checkout_Request_Results - Formatted.xlsx"  # Path to save the output file
highlight_unique_values(input_file, output_file)
